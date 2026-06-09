#!/usr/bin/env python3
"""生成斑马优号结算单 Excel 演示样例（待打款状态）。"""

from __future__ import annotations

import random
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / "结算单样例-待打款-银河号铺.xlsx"
ORDER_COUNT = 32
RNG = random.Random(20260615)

PRODUCTS = [
    "王者荣耀 V10 全皮肤",
    "王者荣耀 国标打野号",
    "原神 60级 满命雷神",
    "原神 55级 多五星号",
    "和平精英 战神号",
    "和平精英 满级枪皮号",
    "英雄联盟 钻石段位",
    "英雄联盟 大师段位",
    "DNF 满级搬砖号",
    "DNF 增幅13武器号",
    "火影忍者 S 忍号",
    "火影忍者 超影段位",
    "崩坏：星穹铁道 满级号",
    "崩坏：星穹铁道 多五星号",
    "永劫无间 修罗段位",
    "逆水寒 高战号",
    "蛋仔派对 全皮肤号",
    "金铲铲之战 大师号",
    "穿越火线 满V号",
    "QQ飞车 永久A车号",
    "第五人格 全角色号",
    "光遇 全图鉴号",
    "明日方舟 高练度号",
    "阴阳师 全图鉴号",
    "梦幻西游 175级号",
    "剑网3 满级号",
    "天涯明月刀 高功力号",
    "妄想山海 满级号",
    "使命召唤手游 满级号",
    "部落冲突 满级号",
    "炉石传说 全卡号",
    "暗黑破坏神不朽 满级号",
]


@dataclass
class OrderLine:
    seq: int
    order_id: str
    product_name: str
    order_status: str
    user_paid: int
    platform_fee: int
    compensation: int
    settle_amount: int


@dataclass
class SettlementDoc:
    settlement_id: str
    withdrawal_id: str
    merchant_id: str
    merchant_name: str
    merchant_type: str
    bank_name: str
    bank_account: str
    bank_holder: str
    created_at: str
    note: str
    status: str
    orders: list[OrderLine]

    @property
    def order_count(self) -> int:
        return len(self.orders)

    @property
    def total_amount(self) -> int:
        return sum(o.settle_amount for o in self.orders)

    @property
    def user_paid_total(self) -> int:
        return sum(o.user_paid for o in self.orders)

    @property
    def fee_total(self) -> int:
        return sum(o.platform_fee for o in self.orders)

    @property
    def comp_total(self) -> int:
        return sum(o.compensation for o in self.orders)


def build_mock_orders(count: int) -> list[OrderLine]:
    orders: list[OrderLine] = []
    for i in range(1, count + 1):
        user_paid = RNG.randint(80, 150) * 100  # 8000-15000
        fee_rate = RNG.choice([0.08, 0.09, 0.10, 0.12])
        platform_fee = round(user_paid * fee_rate)
        compensation = RNG.choice([80, 100, 120, 150, 200, 300, 500])
        settle_amount = user_paid - platform_fee - compensation
        day = RNG.randint(1, 28)
        order_id = f"ORD-202605{day:02d}-{i:03d}"
        product = PRODUCTS[(i - 1) % len(PRODUCTS)]
        orders.append(
            OrderLine(
                seq=i,
                order_id=order_id,
                product_name=product,
                order_status="交易完成",
                user_paid=user_paid,
                platform_fee=platform_fee,
                compensation=compensation,
                settle_amount=settle_amount,
            )
        )
    return orders


def build_settlement() -> SettlementDoc:
    orders = build_mock_orders(ORDER_COUNT)
    return SettlementDoc(
        settlement_id="STL-20260615-001",
        withdrawal_id="WD-20260615-001",
        merchant_id="MCH-002",
        merchant_name="银河号铺",
        merchant_type="号商",
        bank_holder="银河号铺",
        bank_account="6228 **** **** 3366",
        bank_name="工商银行杭州城西支行",
        created_at="2026-06-15 10:20",
        note="卖家申请全额提现，系统自动生成结算单",
        status="未结算（待打款）",
        orders=orders,
    )


def thin_border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def write_kv_row(ws, row: int, label: str, value, label_font, value_font, label_fill=None):
    ws.cell(row=row, column=1, value=label).font = label_font
    if label_fill:
        ws.cell(row=row, column=1).fill = label_fill
    ws.cell(row=row, column=2, value=value).font = value_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)


def build_workbook(doc: SettlementDoc) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "结算单"

    title_font = Font(name="Microsoft YaHei", size=16, bold=True, color="1E3A5F")
    status_font = Font(name="Microsoft YaHei", size=11, bold=True, color="B45309")
    section_font = Font(name="Microsoft YaHei", size=11, bold=True, color="374151")
    label_font = Font(name="Microsoft YaHei", size=10, color="6B7280")
    value_font = Font(name="Microsoft YaHei", size=10, color="111827")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    money_font = Font(name="Microsoft YaHei", size=10, color="111827")
    total_font = Font(name="Microsoft YaHei", size=10, bold=True, color="111827")

    section_fill = PatternFill("solid", fgColor="F3F4F6")
    header_fill = PatternFill("solid", fgColor="2563EB")
    total_fill = PatternFill("solid", fgColor="EFF6FF")
    status_fill = PatternFill("solid", fgColor="FEF3C7")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Row 1: Title + status
    ws.merge_cells("A1:F1")
    ws["A1"] = "斑马优号 · 卖家结算单（应结报表）"
    ws["A1"].font = title_font
    ws["A1"].alignment = left
    ws.merge_cells("G1:H1")
    ws["G1"] = doc.status
    ws["G1"].font = status_font
    ws["G1"].fill = status_fill
    ws["G1"].alignment = center
    ws.row_dimensions[1].height = 28

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="【单头信息】").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    header_rows = [
        ("结算单号", doc.settlement_id),
        ("关联提现单号", doc.withdrawal_id),
        ("主体 ID", doc.merchant_id),
        ("主体名称", doc.merchant_name),
        ("主体类型", doc.merchant_type),
        ("收款户名", doc.bank_holder),
        ("收款账号", doc.bank_account),
        ("开户行", doc.bank_name),
        ("创建时间", doc.created_at),
        ("备注", doc.note),
    ]
    for label, value in header_rows:
        write_kv_row(ws, row, label, value, label_font, value_font)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="【汇总信息】（金额单位：元）").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    summary_rows = [
        ("应结总额", doc.total_amount),
        ("订单笔数", doc.order_count),
        ("用户实付合计", doc.user_paid_total),
        ("平台抽成合计", doc.fee_total),
        ("包赔收入合计", doc.comp_total),
    ]
    for label, value in summary_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = label_font
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = Font(name="Microsoft YaHei", size=10, bold=True, color="111827")
        if isinstance(value, int) and label != "订单笔数":
            c2.number_format = "#,##0"
            c2.alignment = right
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="【订单明细】").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    detail_header_row = row + 1

    headers = [
        "序号",
        "订单号",
        "商品名称",
        "订单状态",
        "用户实付金额",
        "平台抽成金额",
        "包赔金额",
        "结算金额",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=detail_header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border()

    data_start = detail_header_row + 1
    for idx, o in enumerate(doc.orders):
        r = data_start + idx
        values = [
            o.seq,
            o.order_id,
            o.product_name,
            o.order_status,
            o.user_paid,
            o.platform_fee,
            o.compensation,
            o.settle_amount,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border()
            cell.font = money_font
            if col == 1:
                cell.alignment = center
            elif col in (5, 6, 7, 8):
                cell.number_format = "#,##0"
                cell.alignment = right
            else:
                cell.alignment = left

    total_row = data_start + len(doc.orders)
    ws.cell(row=total_row, column=1, value="合计").font = total_font
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=1).fill = total_fill
    for col in range(2, 5):
        ws.cell(row=total_row, column=col, value="").fill = total_fill

    sum_cols = {5: doc.user_paid_total, 6: doc.fee_total, 7: doc.comp_total, 8: doc.total_amount}
    for col, total in sum_cols.items():
        cell = ws.cell(row=total_row, column=col, value=total)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = "#,##0"
        cell.alignment = right
        cell.border = thin_border()

    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = thin_border()

    col_widths = [6, 18, 28, 10, 14, 14, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{data_start}"
    return wb


def verify(doc: SettlementDoc) -> None:
    assert doc.order_count >= 30, "订单数不足 30 条"
    calc_total = sum(o.settle_amount for o in doc.orders)
    assert calc_total == doc.total_amount
    for o in doc.orders:
        assert o.settle_amount == o.user_paid - o.platform_fee - o.compensation
    assert doc.user_paid_total == sum(o.user_paid for o in doc.orders)
    assert doc.fee_total == sum(o.platform_fee for o in doc.orders)
    assert doc.comp_total == sum(o.compensation for o in doc.orders)


def main() -> None:
    doc = build_settlement()
    verify(doc)
    wb = build_workbook(doc)
    wb.save(OUTPUT)
    print(f"已生成: {OUTPUT}")
    print(f"结算单号: {doc.settlement_id}")
    print(f"订单笔数: {doc.order_count}")
    print(f"应结总额: {doc.total_amount:,} 元")
    print(f"用户实付合计: {doc.user_paid_total:,} 元")
    print(f"平台抽成合计: {doc.fee_total:,} 元")
    print(f"包赔收入合计: {doc.comp_total:,} 元")


if __name__ == "__main__":
    main()
